import sys
import os
import tkinter as tk
import threading
from matplotlib import pyplot as plt
import pandas as pd
import numpy as np
import time
import shutil
import tempfile

from openpyxl import load_workbook
from asammdf import MDF, Signal
from tkinter import ttk, filedialog, messagebox
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from python_calamine.pandas import pandas_monkeypatch

"""
Dumarey MF4/Excel Merge and Convert Tool
Last Updated 8/4/2025
Written by Jeffrey Liu

Merges MF4 files with Excel files, outputting a single MF4 file with all signals present and time-aligned.
Additionally, if there is a summary sheet in the Excel file, the values will be extracted and saved as a new Excel file.
This program is designed to accept a queue of jobs to process in the background, requiring minimal user interaction.
One MF4 file can be merged with multiple Excel files or Excel files can be converted to MF4. 
Time-alignment is done using ENGINE SPEED channels. If they are very consistent, LAMBSE channels are used instead.
"""

def resource_path(relative_path):
    # PyInstaller creates a temp folder and stores path in _MEIPASS
    if hasattr(sys, '_MEIPASS'):
        base_path = sys._MEIPASS # type: ignore
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Array to hold jobs for processing
jobs = []

# For debugging and optimization, define a decorator to time functions
def timing(func):
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        print(f"[TIMER] {func.__name__} took {end - start:.3f} seconds")
        return result
    return wrapper

last_mf4_path = "/"
last_excel_path = "/"
# Asks user to select files for input, returns a tuple of (mf4_path, xlsm_path)
def add_job():
    global last_mf4_path, last_excel_path, jobs

    mf4_path = filedialog.askopenfilename(
            title="Select MF4 File (Cancel to create new from Excel)",
            initialdir=last_mf4_path,
            filetypes=[("MF4 files", "*.mf4"), ("All files", "*.*")]
        )
    last_mf4_path = "/".join(mf4_path.split("/")[:-1]) if mf4_path else None
    if last_excel_path == "/":
        last_excel_path = last_mf4_path
    if mf4_path:
        xlsm_paths = filedialog.askopenfilenames(
            title="Select Excel Files",
            initialdir=last_excel_path,
            filetypes=[("Excel Macro-enabled files", "*.xlsm"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        last_excel_path = "/".join(xlsm_paths[0].split("/")[:-1]) if xlsm_paths else None
        jobs.append((mf4_path, list(xlsm_paths)))
    else:
        xlsm_path = filedialog.askopenfilename(
            title="Select Excel File",
            initialdir=last_excel_path,
            filetypes=[("Excel Macro-enabled files", "*.xlsm"), ("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        jobs.append((None, [xlsm_path]))
     
    update_jobs_label()

# Updates the jobs readout to reflect current tasks, is called after adding/removing jobs
def update_jobs_label():
    update_text = "" # Blank string that is appended upon
    for i, (mf4_path, xlsm_paths) in enumerate(jobs):
        update_text += f"Job {i + 1}:\n"
        if mf4_path:
            update_text += f"   MF4: {os.path.basename(mf4_path)}\n"
        for idx, xlsm_path in enumerate(xlsm_paths):
            update_text += f"   Excel {idx+1}: {os.path.basename(xlsm_path)}\n"
    if not update_text:
        update_text = "No tasks queued"
    jobs_label.config(text=update_text) # Update text is applied to the label

# Reads xlsm file, extracts needed columns, and returns a DataFrame with variable names and units
@timing
def read_xlsm_for_merge(xlsm_path, mdf_orig=None):
    # Check for data sheets
    xl = pd.ExcelFile(xlsm_path, engine='calamine')
    if 'Uniplot' in xl.sheet_names:
        sheet_to_read = 'Uniplot'
    elif '10 Hz Data' in xl.sheet_names:
        sheet_to_read = '10 Hz Data'
    else:
        raise ValueError("Neither 'Uniplot' nor '10 Hz Data' sheet found in the Excel file.")

    # Step 1: Read only the header
    header_df = pd.read_excel(xlsm_path, sheet_name=sheet_to_read, nrows=0, engine='calamine')
    excel_columns = list(header_df.columns)

    # Step 2: Compare with MF4 channels if provided 
    if mdf_orig is not None:
        mf4_channels = set(mdf_orig.channels_db.keys())
        needed_columns = [col for col in excel_columns if col not in mf4_channels or col.lower() == 'test time']
    else:
        needed_columns = excel_columns

    # Step 3: Read only needed columns
    df = pd.read_excel(xlsm_path, sheet_name=sheet_to_read, usecols=needed_columns, engine='calamine')
    variable_names = [str(v) if pd.notnull(v) else '' for v in list(df.columns)]
    units = [str(u) if pd.notnull(u) else '' for u in df.iloc[0].tolist()]

    df_data = df.iloc[1:].reset_index(drop=True)
    df_data.columns = variable_names

    if 'Test Time' not in variable_names:
        raise ValueError(f"The '{sheet_to_read}' sheet must contain a 'Test Time' column.")

    # Drops the first x rows of Test Time that are empty
    test_time_idx = df_data.columns.get_loc('Test Time')
    while pd.isnull(df_data.iloc[0, test_time_idx]): #type: ignore
        print(f"Removing empty header row in '{sheet_to_read}' sheet.")
        df_data = df_data.iloc[1:].reset_index(drop=True)
        if df_data.empty:
            raise ValueError(f"The '{sheet_to_read}' sheet does not contain valid data after the header.")

    # Convert all columns except 'Test Time' to numeric
    for col in df_data.columns:
        if col != 'Test Time':
            df_data[col] = pd.to_numeric(df_data[col], errors='coerce')

    # Remove columns that are all NaN (except 'Test Time')
    keep_cols = []
    keep_units = []
    for idx, col in enumerate(variable_names):
        if col == 'Test Time':
            keep_cols.append(col)
            keep_units.append(units[idx])
        elif col in df_data.columns and df_data[col].notna().any():
            keep_cols.append(col)
            keep_units.append(units[idx])
    df_data = df_data[keep_cols]
    print(f'{sheet_to_read} read ', df_data.shape)
    return df_data, keep_cols, keep_units

# Reads the xlsm file and extracts the 'Cumulatives' sheet, saving it as a new Excel file; throws error if the sheet is missing
@timing
def CRS_output(xlsm_paths):
    for sheet in xlsm_paths:
        wb = load_workbook(sheet, read_only=True, data_only=True) # Reads workbook, only data (no macros)
        if 'Cumulatives' in wb.sheetnames:
            ws = wb['Cumulatives']
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            df = pd.DataFrame(data)
        else:
            raise ValueError("The Excel file is missing a 'Cumulatives' sheet, no CRS will be given.")
        output_path = os.path.splitext(sheet)[0] + '_summary.xlsx'
        df.to_excel(output_path, index=False, header=False, engine='openpyxl')

# Brunt of the merging logic
@timing
def merge_xlsm_to_mf4(mf4_path, xlsm_paths, debug_mode=False):
    """
    mf4_path: path to the MF4 file, optional and can be None to instead convert Excel files
    xlsm_paths: list of Excel file paths
    debug_mode: if True, only alignment channels are merged, valuable to save time during testing
    """
    progress['value'] = 0
    mdf = MDF()
    excel_signals_total = 0

    # If mf4_path is provided, files are merged with Excel files
    if mf4_path:
        mdf_orig = MDF(mf4_path)
        # Initially defaults to engine speed channels for alignment
        engspd_channels = [ch for ch in mdf_orig.channels_db if ch.lower().startswith('engine_speed1')]
        if not engspd_channels:
            engspd_channels = [ch for ch in mdf_orig.channels_db if ch.lower().startswith('engine_speed')]

        print(f"Found {len(engspd_channels)} engine speed channels in MF4: {engspd_channels}")

        if not engspd_channels:
            raise ValueError("No engine_speed channel found in MF4.")

        ch_name = engspd_channels[0]
        engspd = mdf_orig.get(ch_name)

        if debug_mode:
            # Only add engine speed channel
            group, index = mdf_orig.channels_db[ch_name][0]
            sig = mdf_orig.get(ch_name, group=group, index=index)
            mdf.append(sig)
        else:
            # Slow! Responsible for 90% of runtime
            # Iterate through all channels in the MF4 file and append them to the MDF object
            for name in mdf_orig.channels_db:
                for group, index in mdf_orig.channels_db[name]:
                    sig = mdf_orig.get(name, group=group, index=index)
                    mdf.append(sig)
                progress.step(50/(len(mdf_orig.channels_db)))
                

        for xlsm_path in xlsm_paths:
            df, variable_names, units = read_xlsm_for_merge(xlsm_path)
            progress.step(20/len(xlsm_paths))
            # Get engine speed column from Excel for alignment
            excel_align_col = next((name for name in variable_names if name.lower() in ['engine speed1', 'engine speed', 'engine_speed']), None)
            if not excel_align_col or excel_align_col not in df.columns:
                raise ValueError(f"Excel file '{xlsm_path}' does not contain a valid engine speed column.")

            mf4_samples = engspd.samples
            mf4_time = engspd.timestamps

            excel_samples = df[excel_align_col].to_numpy(dtype=float)
            excel_time = df['Test Time'].to_numpy(dtype=float)

            # If engine speed has low variance, instead default to lambse_tgt channel for alignment
            if is_signal_consistent(excel_samples):
                lambse_channels = [name for name in variable_names if name.lower().startswith('lambse_tgt')]
                if not lambse_channels:
                    raise ValueError(f"Excel file '{xlsm_path}' does not contain a valid LAMBSE channel for alignment.")
                
                # Check if LAMBSE channel exists in MF4
                mf4_lambse_name = lambse_channels[0]
                if mf4_lambse_name not in mdf_orig.channels_db:
                    raise ValueError(f"Channel '{mf4_lambse_name}' not found in MF4 file for alignment.")
                
                lambse = mdf_orig.get(mf4_lambse_name)
                mf4_samples = lambse.samples
                mf4_time = lambse.timestamps

                group, index = mdf_orig.channels_db[mf4_lambse_name][0]
                sig = mdf_orig.get(mf4_lambse_name, group=group, index=index)
                mdf.append(sig)
                
                excel_align_col = next((name for name in variable_names if name.lower() in ['lambse_tgt[0]']), mf4_lambse_name)
                if excel_align_col not in df.columns:
                    raise ValueError(f"Excel file '{xlsm_path}' does not contain a valid LAMBSE alignment column like '{excel_align_col}'.")
                excel_samples = df[excel_align_col].to_numpy(dtype=float)

            # Determine a common, uniform sample rate
            mf4_rate = 1 / np.mean(np.diff(mf4_time))
            excel_rate = 1 / np.mean(np.diff(excel_time))
            target_rate = min(mf4_rate, excel_rate) # type: ignore
            sample_time = 1 / target_rate
            print(f"Resampling signals to a common rate of {target_rate:.2f} Hz.")

            # Create a new uniform time grid for the longer signal (Excel)
            excel_duration = excel_time[-1] - excel_time[0]
            num_excel_samples_uniform = int(excel_duration * target_rate) + 1
            excel_time_uniform_relative = np.linspace(0, excel_duration, num_excel_samples_uniform)
            excel_samples_uniform = np.interp(excel_time_uniform_relative, excel_time - excel_time[0], excel_samples)

            # Resample the shorter signal (MF4) to the same rate, but keep its original duration
            mf4_duration = mf4_time[-1] - mf4_time[0]
            num_mf4_samples_uniform = int(mf4_duration * target_rate) + 1
            mf4_time_uniform_relative = np.linspace(0, mf4_duration, num_mf4_samples_uniform)
            mf4_samples_uniform = np.interp(mf4_time_uniform_relative, mf4_time - mf4_time[0], mf4_samples)

            # Normalize signals and perform correlation
            excel_norm = (excel_samples_uniform - np.mean(excel_samples_uniform)) / np.std(excel_samples_uniform)
            mf4_norm = (mf4_samples_uniform - np.mean(mf4_samples_uniform)) / np.std(mf4_samples_uniform)
            
            # Correlate the shorter MF4 signal against the longer Excel signal
            correlation = np.correlate(excel_norm, mf4_norm, mode='full')
            correlation = correlation / max(np.abs(correlation))  # Normalize correlation

            # Calculate lag and time offset
            lag_index = np.argmax(correlation)
            # Adjust lag for 'full' correlation mode
            lag_in_samples = lag_index - (len(mf4_norm) - 1)

            time_offset_seconds = lag_in_samples * sample_time
            
            # The offset needed to align Excel's relative time to the MF4's absolute time
            time_offset = mf4_time[0] - (excel_time[0] + time_offset_seconds)
            
            print(f"Detected lag of {time_offset_seconds:.2f} seconds in Excel data.")
            print(f"Calculated time offset: {time_offset:.2f}s")

            aligned_excel_time = excel_time + time_offset

            # Merge signals
            if debug_mode:
                # Only add Excel alignment signal
                unit = units[variable_names.index(excel_align_col)] if excel_align_col in variable_names else ""
                signal = Signal(
                    samples=excel_samples,
                    timestamps=aligned_excel_time,
                    name=excel_align_col if excel_align_col else 'Engine Speed',
                    unit=unit,
                    comment='Alignment Variable'
                )
                mdf.append(signal)
                excel_signals_total += 1
            else:
                # Add all Excel signals except 'Test Time'
                for idx, col in enumerate(variable_names):
                    if col == 'Test Time':
                        continue
                    samples = df[col].to_numpy(dtype=float)
                    unit = units[idx]
                    signal = Signal(
                        samples=samples,
                        timestamps=aligned_excel_time,
                        name=col,
                        unit=unit,
                        comment='Excel signal'
                    )
                    mdf.append(signal)
                    excel_signals_total += 1
                    progress.step(29/(len(variable_names)*len(xlsm_paths)))
    else:
        # No MF4 file, just convert Excel file
        for xlsm_path in xlsm_paths:
            df, variable_names, units = read_xlsm_for_merge(xlsm_path)
            progress.step(20/len(xlsm_paths))
            if 'Test Time' not in variable_names:
                raise ValueError(f"The '{xlsm_path}' file must contain a 'Test Time' column.")
            aligned_excel_time = df['Test Time'].to_numpy(dtype=float)
            for idx, col in enumerate(variable_names):
                if col == 'Test Time':
                    continue
                samples = df[col].to_numpy(dtype=float)
                unit = units[idx]
                signal = Signal(
                    samples=samples,
                    timestamps=aligned_excel_time,
                    name=col,
                    unit=unit,
                    comment='Excel signal'
                )
                mdf.append(signal)
                excel_signals_total += 1
                progress.step(79/(len(variable_names)*len(xlsm_paths)))

    folder = os.path.dirname(xlsm_paths[0])
    base = os.path.splitext(os.path.basename(xlsm_paths[0]))[0]
    if mf4_path:
        output_path = os.path.join(folder, base + "_merged.mf4")
    else:
        output_path = os.path.join(folder, base + "_converted.mf4")
    mdf.save(output_path, compression=1)
    return output_path, excel_signals_total

def remove_job():
    if jobs:
        jobs.pop()
        update_jobs_label()
    else:
        messagebox.showinfo("Info", "No tasks to undo.")

# Tkinter GUI Handling
def threaded_merge(jobs):
    successes = []
    merged_files = []
    errors = []

    def handle_success(output_path, added):
        successes.append((output_path, added))
        merged_files.append(output_path)

    def handle_error(e, xlsm_path):
        errors.append((e, xlsm_path))

    def prepare_progress():
        task_btn_frame.pack_forget()
        convert_btn.pack_forget()
        progress.pack()
    
    def show_results():
        progress.stop()
        progress.pack_forget()
        task_btn_frame.pack(pady=10)
        convert_btn.pack(pady=10)
        results = ""
        if successes:
            results = "Completed Tasks:\n"
            for output_path, added in successes:
                results += f"Added {added} signals to {output_path}\n"

        if errors:
            results += "\nFailed Tasks:\n"
            for error, xlsm_path in errors:
                # ...existing code...
                if isinstance(xlsm_path, (list, tuple)):
                    excel_files = ", ".join(os.path.basename(x) for x in xlsm_path)
                else:
                    excel_files = os.path.basename(xlsm_path)
                results += f"Error: {error}\nExcel File: {excel_files}\n"
                # ...existing code...
        response = messagebox.askyesno("Results", results + "\n\nDo you want to plot the results to verify?")
        if response:
            plot_results()
        jobs_label.config(text="No tasks queued")

    def plot_results():
        plot_window = tk.Toplevel(root)
        plot_window.title("Data Visualization")
        plot_window.geometry("800x600")
        
        # Create a notebook for tabbed plots
        notebook = ttk.Notebook(plot_window)
        notebook.pack(fill='both', expand=True)
        
        for i, file in enumerate(merged_files):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=os.path.basename(file))
            
            mdf = MDF(file)
            fig = plt.figure(figsize=(8, 6))
            ax = fig.add_subplot(111)
            
            if 'Engine Speed' in mdf.channels_db:
                pairs = mdf.channels_db['Engine Speed']
                for group, index in pairs:
                    signal = mdf.get('Engine Speed', group=group, index=index)
                    ax.plot(signal.timestamps, signal.samples, label='Engine Speed')
            if 'ENGINE_SPEED' in mdf.channels_db:
                pairs = mdf.channels_db['ENGINE_SPEED']
                for group, index in pairs:
                    signal = mdf.get('ENGINE_SPEED', group=group, index=index)
                    ax.plot(signal.timestamps, signal.samples, label='ENGINE_SPEED')
                    
            ax.set_title(f"{os.path.basename(file)}")
            ax.set_xlabel("Time (s)")
            ax.set_ylabel("RPM")
            ax.legend()
            
            # Embed the plot in the Tkinter window
            canvas = FigureCanvasTkAgg(fig, master=frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
        
        # This keeps the plot window tied to the main Tkinter window
        plot_window.transient(root)
        plot_window.grab_set()
        root.wait_window(plot_window)

    @timing
    def run_merge():
        jobs_copy = jobs[:]
        for mf4_path, xlsm_paths in jobs_copy:
            try:
                output_path, added = merge_xlsm_to_mf4(mf4_path, xlsm_paths)
                try:
                    CRS_output(xlsm_paths)
                except Exception as e:
                    handle_error(e, xlsm_paths)
                handle_success(output_path, added)
            except Exception as e:
                handle_error(e, xlsm_paths)
            jobs.pop(0)
            root.after(0, update_jobs_label)
        root.after(0, show_results)

    root.after(0, prepare_progress)
    threading.Thread(target=run_merge, daemon=True).start()

# Calculates coefficient of variation (CV) to determine if a signal is consistent
def is_signal_consistent(data):
    cv = np.std(data) / np.mean(data) 
    threshold = 0.01
    if cv < threshold:
        print(cv, "CV is low, signal is consistent")
        return True
    return False

"""
Below is the main GUI setup, using Tkinter and ttk for styling.

main_frame is the container for all widgets.
title_label displays the bolded title text.
desc_label provides subtitle text with brief instructions.
label_frame is a scrollable frame for queued tasks, displaying current jobs.
add_task_btn and undo_task_btn are buttons to add or remove jobs.
convert_btn starts the merge and convert process.
"""
root = tk.Tk()
root.title("MF4 + Excel Merge and Convert")
root.geometry("400x600")
root.resizable(False, True)

icon_path = resource_path(os.path.join("xlsx_mdf_converter", "icons", "dumarey_favicon.ico"))

# Gets the icon for the application, handles PyInstaller packaging
if hasattr(sys, '_MEIPASS'):
    temp_icon = tempfile.NamedTemporaryFile(delete=False, suffix='.ico')
    shutil.copyfile(icon_path, temp_icon.name)
    root.iconbitmap(temp_icon.name)
else:
    root.iconbitmap(icon_path)

style = ttk.Style(root)
style.theme_use('default')
green = "#28a745"
style.configure("Green.TButton", foreground="white", background=green)
style.map("Green.TButton",
        background=[('active', '#218838'), ('!active', green)])
red = "#972e26"
style.configure("Red.TButton", foreground="white", background=red)
style.map("Red.TButton",
        background=[('active', '#c82333'), ('!active', red)])

main_frame = ttk.Frame(root, padding=20, style="Main.TFrame")
main_frame.pack(expand=True, fill='both')

title_label = ttk.Label(main_frame, text="MF4/Excel Merge and Convert", font=("Segoe UI", 16, "bold"))
title_label.pack(pady=(0, 10))

desc_label = ttk.Label(
    main_frame,
    text="Select an MF4 and an Excel file to merge.\n"
        "Or cancel MF4 selection to convert Excel to MF4.",
    font=("Segoe UI", 10),
    style="Main.TLabel"
)
desc_label.pack(pady=(0, 20))

label_frame = ttk.LabelFrame(main_frame, text="Queued Tasks")
label_frame.configure(style="White.TLabelframe")
style.configure("White.TLabelframe", background="#d9d9d9")
style.configure("White.TLabelframe.Label", background="#d9d9d9")

label_frame.pack(fill='both', expand=True, pady=(0, 10), padx=20)

canvas = tk.Canvas(label_frame, borderwidth=0, highlightthickness=0, height=120, bg="#d9d9d9")
scrollbar = ttk.Scrollbar(label_frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas, bg="#d9d9d9")  # Use tk.Frame for easier bg control

def on_frame_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

scrollable_frame.bind("<Configure>", on_frame_configure)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

jobs_label = tk.Label(
    scrollable_frame,
    text="No tasks queued",
    font=("Segoe UI", 9),
    fg="black",
    bg="#d9d9d9",
    justify="left",
    anchor="nw"
)
jobs_label.pack(pady=(5, 0), anchor="w", fill="x")

# Implements scrolling behavior of the label frame
def _on_mousewheel(event):
    # Get current scroll position (returns a tuple of fractions, e.g., (0.0, 1.0) means fully scrolled)
    first, last = canvas.yview()
    delta = int(-1*(event.delta/120))
    if delta < 0 and first <= 0.0:
        return  
    if delta > 0 and last >= 1.0:
        return
    canvas.yview_scroll(delta, "units")

canvas.bind_all("<MouseWheel>", _on_mousewheel)
task_btn_frame = ttk.Frame(main_frame)
task_btn_frame.pack(pady=10)

add_task_btn = ttk.Button(task_btn_frame, text="Add Task", command=lambda: add_job())
add_task_btn.configure(width=16)
add_task_btn.pack(side='left', padx=(0, 5))

undo_task_btn = ttk.Button(task_btn_frame, text="Undo", command=lambda: remove_job())
undo_task_btn.configure(style="Red.TButton")
undo_task_btn.pack(side='left')

convert_btn = ttk.Button(main_frame, text="Merge and Convert", command=lambda: threaded_merge(jobs), style="Green.TButton")
convert_btn.pack(pady=10)

progress = ttk.Progressbar(main_frame, mode='determinate', length=250)
progress.pack(pady=10)
progress.pack_forget()

status_var = tk.StringVar()
status_label = ttk.Label(main_frame, textvariable=status_var, font=("Segoe UI", 9), foreground="gray")
status_label.pack(pady=(5, 0))

def on_closing():
    """Handles the window closing event to prevent shutdown errors."""
    plt.close('all') 
    root.quit()       
    root.destroy()    

root.protocol("WM_DELETE_WINDOW", on_closing)
root.mainloop()